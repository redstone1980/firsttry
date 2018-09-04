IDCHECK = '*IDN?'
MODE_LTE = 'INST:SEL LTE'
MODE_LTEAFDD = 'INST:SEL LTEAFDD'
BW1_4 ='B1M4'; BW3 = 'B3M'; BW5 = 'B5M'; BW10 = 'B10M'; BW15 = 'B15M'; BW20 = 'B20M'
BOOKADD = 'test report for 4150 in LR18_1.xlsx'
INS_ID = 3
FREQ = 1937.5
OFFSET = -44.58
ATT = 20
import visa
import time
from openpyxl import load_workbook
from openpyxl import workbook
from tkinter import *
from tkinter import ttk
rm = visa.ResourceManager()
list = rm.list_resources()

class CHPOWER_measure:
	def __init__(self,command,index,mode,bw,freq):
		self.command = command
		self.index = index
		self.mode = mode
		self.bw = 'RAD:STAN:PRES '+bw
		self.freq = 'FREQ:CENT '+str(freq)+' MHz'
	def IDtest(self):
		print(list[self.index])
		n9020a = rm.open_resource(list[self.index])
		x = n9020a.query(self.command)
		print(x)
	def template_init(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('CONF:CHP')
		n9020a.write('INIT:CONT ON')
		n9020a.write('INIT:CHP')
		n9020a.write(self.bw)
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('READ:CHP?')
		time.sleep(1)
		alldata = n9020a.read_ascii_values()
		POWdata = alldata[0]
		print('CHPower = '+str(POWdata) + ' dBm')
		return POWdata
	def ContON(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write('INIT:CONT ON')


class OBW_measure:
	def __init__(self,command,index,mode,bw,freq):
		self.command = command
		self.index = index
		self.mode = mode
		self.bw = 'RAD:STAN:PRES '+bw
		self.freq = 'FREQ:CENT '+str(freq)+' MHz'
	def IDtest(self):
		print(list)
		n9020a = rm.open_resource(list[self.index])
		x = n9020a.query(self.command)
		print(x)
	def template_init(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('CONF:OBW')
		n9020a.write('INIT:CONT ON')
		n9020a.write('INIT:OBW')
		n9020a.write(self.bw)
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('READ:OBW?')
		time.sleep(1)
		alldata = n9020a.read_ascii_values()
		OBWdata = alldata[0]
		print('OBW = '+str(OBWdata/1e6)+' MHz')
		return OBWdata/1e6


class ACLR_measure:
	def __init__(self,command,index,mode,bw,freq):
		self.command = command
		self.index = index
		self.mode = mode
		self.bw = 'RAD:STAN:PRES '+bw
		self.freq = 'FREQ:CENT '+str(freq)+' MHz'
	def IDtest(self):
		print(list[self.index])
		n9020a = rm.open_resource(list[self.index])
		x = n9020a.query(self.command)
		print(x)
	def ACLR3_84(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('CONF:ACP')
		n9020a.write('ACP:BAND:AUTO ON')
		n9020a.write('ACP:BWID:VID:AUTO ON')
		n9020a.write('INIT:CONT ON')
		n9020a.write('INIT:ACP')
		n9020a.write('ACP:OFFS:LIST:BAND 3.84MHz,3.84MHz,3.84MHz,3.84MHz,3.84MHz,3.84MHz')
		n9020a.write('ACP:OFFS:LIST 10MHz,15MHz,0,0,0,0')
#		n9020a.write('ACP:OFFS:LIST:STA 1,1,0,0,0,0')
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('READ:ACP?')
		time.sleep(5)
		alldata = n9020a.read_ascii_values()
		ACLR3_84data = [alldata[4],alldata[6],alldata[8],alldata[10]]
		print('ACLR3_84')
		print(ACLR3_84data)
		return ACLR3_84data		
	def template_init(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('CONF:ACP')
		n9020a.write('ACP:BAND:AUTO ON')
		n9020a.write('ACP:BWID:VID:AUTO ON')
		n9020a.write('INIT:CONT ON')
		n9020a.write('INIT:ACP')
		n9020a.write(self.bw)
		n9020a.write('ACP:OFFS:LIST:BAND 15MHz,15MHz,15MHz,15MHz,15MHz,15MHz')
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('READ:ACP?')
		time.sleep(5)
		alldata = n9020a.read_ascii_values()
		ACLRdata = [alldata[4],alldata[6],alldata[8],alldata[10]]
		print('ACLR')
		print(ACLRdata)
		return ACLRdata

class EVM_measure:
	def __init__(self,command,index,mode,bw,freq):
		self.command = command
		self.index = index
		self.mode = mode
		self.bw = 'RAD:STAN:PRES '+bw
		self.freq = 'FREQ:CENT '+str(freq)+' MHz'
		self.EVMdata = {'EVM':0,'DLRS':0,'OFDM_TX_Power':0,'FreqErr':0}
		self.n9020a = rm.open_resource(list[self.index])
	def IDtest(self):
		print(list[self.index])
		n9020a = rm.open_resource(list[self.index])
		x = n9020a.query(self.command)
		print(x)
		print('this is EVM_measure')
	def template_init(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('POW:ATT 20')
		n9020a.write('CONF:EVM')
#		n9020a.write('TRIG:EXT1 LEV 0.5V')
		n9020a.write('TRIG:EVM:SOUR EXT1')
		n9020a.write('INIT:EVM')
		n9020a.write('INIT:CONT ON')
		n9020a.write(self.bw)
		time.sleep(10)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('FETC:EVM1?')
		time.sleep(5)
		alldata = n9020a.read_ascii_values()
		self.EVMdata['EVM'] = alldata[0]
		self.EVMdata['DLRS'] = alldata[10]
		self.EVMdata['OFDM_TX_Power'] = alldata[11]
		self.EVMdata['FreqErr'] = alldata[12]
		print(self.EVMdata)
#		print(alldata)
		return self.EVMdata
		#print(alldata)
	def gotoCEVM(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write('INIT:CONT ON')
		n9020a.write('CONF:CEVM')
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')



class SEM_measure:
	def __init__(self,command,index,mode,bw,freq):
		self.command = command
		self.index = index
		self.mode = mode
		self.bw = 'RAD:STAN:PRES '+bw
		self.freq = 'FREQ:CENT '+str(freq)+' MHz'
	def IDtest(self):
		print(list[self.index])
		n9020a = rm.open_resource(list[self.index])
		x = n9020a.query(self.command)
		print(x)
	def template_init(self):
		n9020a = rm.open_resource(list[self.index])
		n9020a.write(self.mode)
		n9020a.write(self.freq)
		n9020a.write('CONF:SEM')
		n9020a.write('INIT:CONT ON')
		n9020a.write('INIT:SEM')
		n9020a.write(self.bw)
		n9020a.write('SEM:OFFS1:LIST:STAT 1,1,1,1')
#		n9020a.write('SEM:FREQ:SPAN 15MHz')
		n9020a.write('SEM:OFFS:LIST:BAND 51 KHz, 100KHz, 1MHz, 100KHz')
		n9020a.write('SEM:OFFS1:LIST:FREQ:STAR  50 kHz, 5.05 MHz, 10.05 MHz, 0.05MHz')
		n9020a.write('SEM:OFFS1:LIST:FREQ:STOP 5.05 MHz, 10.05 MHz, 15 MHz, 0.95MHz')
		n9020a.write('SEM:OFFS1:LIST:STOP:ABS  -14 dBm,-14 dBm,-15 dBm,-15 dBm')
		n9020a.write('SEM:OFFS1:LIST:START:ABS -7 dBm,-14 dBm,-15 dBm,-15 dBm')
		time.sleep(5)
		n9020a.write('INIT:CONT OFF')
		n9020a.write('READ:SEM?')
		time.sleep(5)
		alldata = n9020a.read_ascii_values()
		SEMdata = [[0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0]]
		SEMdata[0][0] = alldata[13]; SEMdata[0][1] = alldata[18]; SEMdata[0][2] = alldata[23]; SEMdata[0][3] = alldata[28]; SEMdata[0][4] = alldata[33]; SEMdata[0][5] = alldata[38]; SEMdata[0][6] = alldata[43]; SEMdata[0][7] = alldata[48]
		SEMdata[1] = alldata[-12:-4]
		print('SEM')
		print(SEMdata)
		return SEMdata

class DATA_record:
	def __init__(self,dir,sheetname):
		self.dir = dir
		self.book = load_workbook(dir)
		if sheetname in self.book.sheetnames:
			self.sheet = self.book[sheetname]
		else:
			self.sheet = self.book.create_sheet()
			self.sheet.title = sheetname
		self.book.save(self.dir)
	def write(self,row,column,data):
		self.sheet.cell(row = row, column = column).value = data
#		print(cell+' write :'+str(self.sheet[cell].value))
	def read(self,row,column):
	  	print(' read :'+str(self.sheet.cell(row = row, column = column).value))
	def writeline(self,data):
		self.sheet.append(data)
	def save(self):
		self.book.save(self.dir)

def getrealtime():

	timeformat = '%Y-%m-%d %X'
	realtime = time.strftime(timeformat,time.localtime(time.time()))
	return realtime
"""
aclr2 = OBW_measure(IDCHECK,3,MODE_LTE,BW15,2000)
aclr2.IDtest()
OBWdata = aclr2.template_init()/1e6

datarecord = DATA_record(BOOKADD,'OBW')
datarecord.write('D8',OBWdata)
datarecord.save()

datarecord2 = DATA_record(BOOKADD,'OBW')
datarecord2.read('D8')
"""

#aclr2 = SEM_measure(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ)
#aclr2.IDtest()
#semdata = aclr2.template_init()
#print(semdata)


#aclr2 = ACLR_measure(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ)
#aclr2.IDtest()
#aclr2.template_init()

#aclr2 = EVM_measure(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ)
#aclr2.IDtest()
#aclr2.template_init()

"""
wb = load_workbook('test report for 4150 in LR18_1.xlsx')
ws = wb['OBW']
ws['D8'] = OBWdata/1e6
print(ws['D8'].value)
wb.save('test report for 4150 in LR18_1.xlsx')
"""
def ETM1_1_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 1.1*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

	chpower = CHPOWER_measure(command,instID,mode,bwstd,freq)
	powerdata = chpower.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_CHpower.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	chpower.ContON()
#	print(screen_com)
#	print(powerdata)
#	datarecord = DATA_record(BOOKADD,'Output Power')
#	datarecord.write('F8',powerdata)
#	datarecord.read('F8')
#	datarecord.save()
	time.sleep(1)

	obw = OBW_measure(command,instID,mode,bwstd,freq)
	obwdata = obw.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_OBW.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)
#	print(obwdata)
#	datarecord = DATA_record(BOOKADD,'OBW')
#	datarecord.write('D8',obwdata)
#	datarecord.read('D8')
#	datarecord.save()
	time.sleep(1)
	
	evm = EVM_measure(command,instID,mode,bwstd,freq)
	dlrsdata = evm.template_init()['DLRS']
#	evm.gotoCEVM()
	time.sleep(3)
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_DLRS.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)
	time.sleep(1)
#	print(dlrsdata)
#	datarecord = DATA_record(BOOKADD,'DL RS power')
#	datarecord.write('D8',dlrsdata)
#	datarecord.read('D8')
#	datarecord.save()
	time.sleep(1)
	
	aclr = ACLR_measure(command,instID,mode,bwstd,freq)
	aclrdata = aclr.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_ACLR.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)
	time.sleep(3)

	aclr3_84data = aclr.ACLR3_84()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_ACLR3_84.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)
#	print(aclrdata)
#	datarecord = DATA_record(BOOKADD,'ACLR')
#	datarecord.write('E8',aclrdata[0]); datarecord.write('F8',aclrdata[2]); datarecord.write('G8',aclrdata[4]); datarecord.write('H8',aclrdata[6]);
#	datarecord.read('E8')
#	datarecord.save()
	time.sleep(3)
	
	sem = SEM_measure(command,instID,mode,bwstd,freq)
	semdata = sem.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.1_SEM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
#	print(screen_com)
	n9020a.write(screen_com)
#	print(semdata)
#	alldata = [powerdata,obwdata,dlrsdata,aclrdata,semdata]
#	testcase = ['Channel Power','OBW','DL RS Power','ACLR','SEM']
	datarecord = DATA_record(BOOKADD,'Raw Data')
#	rowlength = len(list(datarecord.sheet.rows))
	datarecord.writeline([getrealtime(),])
	
	datarecord.writeline([TXchannel,freq,'ETM1.1','Channel Power',powerdata])
	datarecord.writeline([TXchannel,freq,'ETM1.1','OBW',obwdata])
	datarecord.writeline([TXchannel,freq,'ETM1.1','DL RS Power',dlrsdata])
	datarecord.writeline([TXchannel,freq,'ETM1.1','ACLR',aclrdata[0],aclrdata[1],aclrdata[2],aclrdata[3]])
	datarecord.writeline([TXchannel,freq,'ETM1.1','ACLR3_84',aclr3_84data[0],aclr3_84data[1],aclr3_84data[2],aclr3_84data[3]])
	datarecord.writeline([TXchannel,freq,'ETM1.1','SEM',semdata[0][0],semdata[0][1],semdata[0][2],semdata[0][3],semdata[0][4],semdata[0][5],semdata[0][6],semdata[0][7]])
	datarecord.writeline(['','','','',semdata[1][0],semdata[1][1],semdata[1][2],semdata[1][3],semdata[1][4],semdata[1][5],semdata[1][6],semdata[1][7]])
	datarecord.save()
	time.sleep(1)
	print('\n')

def ETM1_2_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 1.2*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

	aclr = ACLR_measure(command,instID,mode,bwstd,freq)
	aclrdata = aclr.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.2_ACLR.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)

	aclr3_84data = aclr.ACLR3_84()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.2_ACLR3_84.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
#	print(screen_com)
#	print(aclrdata)
#	datarecord = DATA_record(BOOKADD,'ACLR')
#	datarecord.write('E9',aclrdata[0]); datarecord.write('F9',aclrdata[2]); datarecord.write('G9',aclrdata[4]); datarecord.write('H9',aclrdata[6]);
#	datarecord.read('E9')
#	datarecord.save()
	
	time.sleep(1)
	
	sem = SEM_measure(command,instID,mode,bwstd,freq)
	semdata = sem.template_init()
	time.sleep(3)
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM1.2_SEM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	n9020a.write(screen_com)

	datarecord = DATA_record(BOOKADD,'Raw Data')
	datarecord.writeline([getrealtime(),])
	datarecord.writeline([TXchannel,freq,'ETM1.2','ACLR',aclrdata[0],aclrdata[1],aclrdata[2],aclrdata[3]])
	datarecord.writeline([TXchannel,freq,'ETM1.2','ACLR3_84',aclr3_84data[0],aclr3_84data[1],aclr3_84data[2],aclr3_84data[3]])
	datarecord.writeline([TXchannel,freq,'ETM1.2','SEM',semdata[0][0],semdata[0][1],semdata[0][2],semdata[0][3],semdata[0][4],semdata[0][5],semdata[0][6],semdata[0][7]])
	datarecord.writeline(['','','','',semdata[1][0],semdata[1][1],semdata[1][2],semdata[1][3],semdata[1][4],semdata[1][5],semdata[1][6],semdata[1][7]])
	datarecord.save()
	time.sleep(1)
	print('\n')

def ETM2_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 2*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

#	chpower = CHPOWER_measure(command,instID,mode,bwstd,freq)
#	powerdata = chpower.template_init()
#	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM2_CHpower.png'
#	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
#	time.sleep(3)
#	n9020a.write(screen_com)

#	print(powerdata)
#	datarecord = DATA_record(BOOKADD,'Total power dynamic range')
#	datarecord.write('E9',powerdata)
#	datarecord.read('E9')
#	datarecord.save()
	time.sleep(1)

	evm = EVM_measure(command,instID,mode,bwstd,freq)
	data = evm.template_init()
	evm.gotoCEVM()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM2_EVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
	evmdata = data['EVM']
	freqerrdata = data['FreqErr']
	ofdmpowerdata = data['OFDM_TX_Power']
#	print(dlrsdata)
#	datarecord = DATA_record(BOOKADD,'EVM')
#	datarecord.write('E7',evmdata)
#	datarecord.read('E7')
#	datarecord.save()

#	datarecord = DATA_record(BOOKADD,'Freq error')
#	datarecord.write('E8',freqerrdata)
#	datarecord.read('E8')
	datarecord = DATA_record(BOOKADD,'Raw Data')
	datarecord.writeline([getrealtime(),])
	datarecord.writeline([TXchannel,freq,'ETM2','OFDM_TX_Power',ofdmpowerdata])
	datarecord.writeline([TXchannel,freq,'ETM2','EVM',evmdata])
	datarecord.writeline([TXchannel,freq,'ETM2','FreqErr',freqerrdata])	
	datarecord.save()
	time.sleep(1)
	print('\n')

def ETM3_1_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 3.1*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

#	chpower = CHPOWER_measure(command,instID,mode,bwstd,freq)
#	powerdata = chpower.template_init()
#	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_1_CHpower.png'
#	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
#	time.sleep(3)
#	n9020a.write(screen_com)
#	print(powerdata)
#	datarecord = DATA_record(BOOKADD,'Total power dynamic range')
#	datarecord.write('D9',powerdata)
#	datarecord.read('D9')
#	datarecord.save()
	time.sleep(1)

	evm = EVM_measure(command,instID,mode,bwstd,freq)
	data = evm.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_1_EVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)

	evm.gotoCEVM()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_1_CEVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)
	evmdata = data['EVM']
	freqerrdata = data['FreqErr']
	ofdmpowerdata = data['OFDM_TX_Power']
#	print(dlrsdata)
#	datarecord = DATA_record(BOOKADD,'EVM')
#	datarecord.write('E10',evmdata)
#	datarecord.read('E10')
#	datarecord.save()

#	datarecord = DATA_record(BOOKADD,'Freq error')
#	datarecord.write('E11',freqerrdata)
#	datarecord.read('E11')
	datarecord = DATA_record(BOOKADD,'Raw Data')
	datarecord.writeline([getrealtime(),])
	datarecord.writeline([TXchannel,freq,'ETM3.1','OFDM_TX_Power',ofdmpowerdata])
	datarecord.writeline([TXchannel,freq,'ETM3.1','EVM',evmdata])
	datarecord.writeline([TXchannel,freq,'ETM3.1','FreqErr',freqerrdata])	
	datarecord.save()
	time.sleep(1)
	print('\n')

def ETM3_2_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 3.2*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

	evm = EVM_measure(command,instID,mode,bwstd,freq)
	data = evm.template_init()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_2_EVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)

	evm.gotoCEVM()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_2_CEVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)

	evmdata = data['EVM']
	freqerrdata = data['FreqErr']
#	print(dlrsdata)
#	datarecord = DATA_record(BOOKADD,'EVM')
#	datarecord.write('E13',evmdata)
#	datarecord.read('E13')
#	datarecord.save()

#	datarecord = DATA_record(BOOKADD,'Freq error')
#	datarecord.write('E14',freqerrdata)
#	datarecord.read('E14')
	datarecord = DATA_record(BOOKADD,'Raw Data')
	datarecord.writeline([getrealtime(),])
	datarecord.writeline([TXchannel,freq,'ETM3.2','EVM',evmdata])
	datarecord.writeline([TXchannel,freq,'ETM3.2','FreqErr',freqerrdata])	
	datarecord.save()
	time.sleep(1)
	print('\n')

def ETM3_3_test(command,instID,mode,bwstd,freq,offset,TXchannel):
	print('*********ETM 3.3*********')
	n9020a = rm.open_resource(list[instID])
	n9020a.write('CORR:BTS:GAIN '+str(offset))

	time.sleep(1)

	evm = EVM_measure(command,instID,mode,bwstd,freq)
	data = evm.template_init()
	evm.gotoCEVM()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_3_EVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)

	evm.gotoCEVM()
	pic_name = TXchannel+'_'+str(freq)+'_'+bwstd+'_ETM3_3_CEVM.png'
	screen_com = 'MMEM:STOR:SCR '+'"'+pic_name+'"'
	time.sleep(3)
	n9020a.write(screen_com)

	evmdata = data['EVM']
	freqerrdata = data['FreqErr']
#	print(dlrsdata)
#	datarecord = DATA_record(BOOKADD,'EVM')
#	datarecord.write('E16',evmdata)
#	datarecord.read('E16')

#	datarecord = DATA_record(BOOKADD,'Freq error')
#	datarecord.write('E17',freqerrdata)
#	datarecord.read('E17')
	datarecord = DATA_record(BOOKADD,'Raw Data')
	datarecord.writeline([getrealtime(),])
	datarecord.writeline([TXchannel,freq,'ETM3.3','EVM',evmdata])
	datarecord.writeline([TXchannel,freq,'ETM3.3','FreqErr',freqerrdata])	
	datarecord.save()
	time.sleep(1)
	print('\n')

def main():
	for i in range(1):
		print('***********This is NO:'+str(i)+' test***********')
		ETM1_1_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
#		ETM1_2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
#		ETM2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
#		ETM3_1_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
#		ETM3_2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
#		ETM3_3_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,FREQ,OFFSET,'TX1')
		print('\n')

def teststart(event,freq,model,channel,offset):
#	print(Emodel['values'][Emodel.current()])
	print('*************TEST START!!!!!!**************')
	print('*************Will wait 30 seconds!**************')	
	for i in range(6):
		time.sleep(5)
		print(30-i*5)
	if model == 'ETM1.1':
		ETM1_1_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	elif model =='ETM1.2':
		ETM1_2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	elif model =='ETM2':
		ETM2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	elif model =='ETM3.1':
		ETM3_1_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	elif model =='ETM3.2':
		ETM3_2_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	elif model =='ETM3.3':
		ETM3_3_test(IDCHECK,INS_ID,MODE_LTEAFDD,BW15,freq,offset,channel)
	print('*************TEST DONE!!!!!!**************')
root = Tk()

Lfreq = ttk.Label(root,text = 'Freq:      ')
Lfreq.grid(row=0,column=0)
Efreq = ttk.Combobox(root)
Efreq['values'] = (1937.5,1955,1970,1987.5)
Efreq.grid(row=0,column=1)

Lmodel = ttk.Label(root,text = 'Test_Model:')
Lmodel.grid(row=1,column=0)
Emodel = ttk.Combobox(root)
Emodel['values'] = ('ETM1.1','ETM1.2','ETM2','ETM3.1','ETM3.2','ETM3.3')
Emodel.grid(row=1,column=1)

Lchannel = ttk.Label(root,text = 'Channel_No:')
Lchannel.grid(row=2,column=0)
Echannel = ttk.Combobox(root)
Echannel['values'] = ('TX1','TX2','TX3','TX4')
Echannel.grid(row=2,column=1)

Loffset = ttk.Label(root,text = 'Offset:    ')
Loffset.grid(row=3,column=0)
Eoffset = ttk.Combobox(root)
Eoffset['values'] = (-44.58,-44.85)
Eoffset.grid(row=3,column=1)

Btest = ttk.Button(root,text = 'START TEST')

Btest.bind('<Button-1>',lambda x:teststart(x,Efreq['values'][Efreq.current()],Emodel['values'][Emodel.current()],Echannel['values'][Echannel.current()],Eoffset['values'][Eoffset.current()]))

Btest.grid(row=4,column=0)

root.mainloop()
#if __name__ == '__main__':
#	main()

